import sys
import cmd
import re
import textwrap
import os.path

from openpyxl import load_workbook
from jinja2 import Template
from version import __version__ as VERSION

class Xlsx2Wiki(cmd.Cmd):
    """
    """

    def __init__(self, file_path = None):
        cmd.Cmd.__init__(self)

        self.file_path = file_path
        self.workbook = load_workbook(filename = self.file_path)

    def transform(self, sheet_name, target_range, tmpl_path):
        ws = self.workbook.get_sheet_by_name(sheet_name.decode('utf-8'))
        with open(tmpl_path) as f:
            tmpl = f.read()
            rendered = Template(textwrap.dedent(tmpl).strip()).render({
                'rows': ws[target_range],
            })
            print(rendered.encode('utf-8'))

def run_(args):
    if args is None:
        args = sys.argv[1:]
    if args:
        if args[0] in ('-h', '--help', 'help'):
            print('usage: {} file-name sheet-name transform-range template'.format(sys.argv[0]))
            return 0
        elif args[0] in ('-v', '--version'):
            print('{} {}'.format(os.path.basename(sys.argv[0]), VERSION))
            return 0
        else:
            file_path = args[0]
            sheet_name = args[1]
            target_range = args[2]
            tmpl_path = args[3]
            xlsx2wiki = Xlsx2Wiki(file_path)
            xlsx2wiki.transform(sheet_name, target_range, tmpl_path)

def run(args=None):
    """
    Main entry point.
    """
    return run_(args)

if __name__ == '__main__':
    sys.exit(run())
